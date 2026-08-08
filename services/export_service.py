"""현재 매물 목록을 사무실 내부용 엑셀 파일로 만드는 기능."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from services.lot_address_service import split_lot_address
from services.record_number import consultation_number, contract_number, listing_number


EXPORT_COLUMNS = [
    ("매물번호", "listing_number", "text"),
    ("접수일", "received_date", "date"),
    ("매물 상태", "listing_status", "text"),
    ("매물 보유처", "listing_holder", "text"),
    ("건물명", "building_name", "text"),
    ("지번 지역", "lot_area", "text"),
    ("번지 번호", "lot_number", "text"),
    ("호실", "unit_number", "text"),
    ("층", "floor_number", "number"),
    ("룸 형태", "room_type", "text"),
    ("방향", "direction", "text"),
    ("보증금(만원)", "deposit_manwon", "number"),
    ("월세(만원)", "monthly_rent_manwon", "number"),
    ("관리비(만원)", "management_fee_manwon", "number"),
    ("관리비 메모", "management_fee_note", "text"),
    ("입주 가능", "availability_type", "text"),
    ("입주 가능일", "available_from_date", "date"),
    ("퇴실 예정일", "move_out_due_date", "date"),
    ("사진 보유 여부", "has_listing_photos", "text"),
    ("재확인 예정일", "next_check_date", "date"),
    ("공동현관 비밀번호", "common_entrance_password", "text"),
    ("방문 방법", "access_method", "text"),
    ("방문 비밀번호", "unit_access_password", "text"),
    ("엘리베이터", "has_elevator", "text"),
    ("주차", "parking_status", "text"),
    ("호실 옵션", "unit_options", "text"),
    ("건물 내부 메모", "building_internal_note", "text"),
    ("호실 내부 메모", "unit_internal_note", "text"),
    ("매물 메모", "listing_note", "text"),
    ("이번 매물 옵션 변경 메모", "option_change_note", "text"),
]

CONTRACT_EXPORT_COLUMNS = [
    ("계약번호", "contract_number", "text"),
    ("매물번호", "listing_number", "text"),
    ("건물명", "building_name", "text"),
    ("지번주소", "lot_address", "text"),
    ("호실", "unit_number", "text"),
    ("매물 접수일", "received_date", "date"),
    ("계약 유형", "contract_type", "text"),
    ("계약 상태", "contract_status", "text"),
    ("계약 진행 시작일", "contract_progress_date", "date"),
    ("정식 계약일", "formal_contract_date", "date"),
    ("임대차 시작일", "contract_start_date", "date"),
    ("임대차 종료일", "contract_end_date", "date"),
    ("임대차 기간(개월)", "term_months", "number"),
    ("계약금 전체(만원)", "contract_deposit_manwon", "number"),
    ("가계약금 수령액(만원)", "provisional_deposit_manwon", "number"),
    ("계약금 추가 수령 예정일", "remaining_deposit_due_date", "date"),
    ("잔금(만원)", "balance_manwon", "number"),
    ("잔금 예정일", "balance_due_date", "date"),
    ("계약 메모", "contract_note", "text"),
]

CONSULTATION_EXPORT_COLUMNS = [
    ("상담번호", "consultation_number", "text"),
    ("연결 매물번호", "listing_number", "text"),
    ("건물명", "building_name", "text"),
    ("지번주소", "lot_address", "text"),
    ("호실", "unit_number", "text"),
    ("매물 접수일", "received_date", "date"),
    ("상담 구분", "consultation_category", "text"),
    ("고객 연락처", "customer_phone", "text"),
    ("상담일", "consulted_date", "date"),
    ("상담 종류", "consultation_type", "text"),
    ("유입 경로", "consultation_source", "text"),
    ("희망 지역", "desired_area", "text"),
    ("희망 룸 형태", "desired_room_type", "text"),
    ("희망 보증금(만원)", "desired_deposit_manwon", "number"),
    ("희망 월세(만원)", "desired_monthly_rent_manwon", "number"),
    ("희망 입주 가능일", "desired_available_from_date", "date"),
    ("다음 연락일", "next_contact_date", "date"),
    ("상담 상태", "consultation_status", "text"),
]

TODAY_TASK_EXPORT_COLUMNS = [
    ("구분", "구분", "text"),
    ("업무 구분", "업무 구분", "text"),
    ("업무번호", "업무번호", "text"),
    ("연결 매물번호", "연결 매물번호", "text"),
    ("해야 할 일", "해야 할 일", "text"),
    ("기한", "기한", "text"),
    ("건물명", "건물명", "text"),
    ("지번", "지번", "text"),
    ("호실", "호실", "text"),
    ("상태", "상태", "text"),
]


def _excel_value(value: Any, value_type: str) -> Any:
    if value in (None, ""):
        return None
    if value_type == "date" and isinstance(value, str):
        return date.fromisoformat(value)
    return value


def _create_excel(rows: list[dict[str, Any]], columns: list[tuple[str, str, str]], title: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append([label for label, _, _ in columns])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for row in rows:
        sheet.append([_excel_value(row.get(key), value_type) for _, key, value_type in columns])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, (_, _, value_type) in enumerate(columns, start=1):
        column = get_column_letter(index)
        sheet.column_dimensions[column].width = 14 if value_type != "text" else 20
        if value_type == "date":
            for cell in sheet[column][1:]:
                cell.number_format = "yyyy-mm-dd"
        elif value_type == "number":
            for cell in sheet[column][1:]:
                cell.number_format = '#,##0'

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def create_current_listing_excel(rows: list[dict[str, Any]]) -> bytes:
    """명시한 열만 사용해 현재 매물 내부 업무용 엑셀을 만든다. 연락처 관련 열은 포함하지 않는다."""
    export_rows = []
    for row in rows:
        lot_area, lot_number = split_lot_address(row.get("lot_address"))
        export_rows.append({**row, "listing_number": listing_number(row.get("listing_id")), "lot_area": lot_area, "lot_number": lot_number})
    return _create_excel(export_rows, EXPORT_COLUMNS, "현재 매물")


def create_contract_excel(rows: list[dict[str, Any]]) -> bytes:
    """계약 조회 결과를 내보낸다. 계약자 연락처는 포함하지 않는다."""
    export_rows = [{**row, "contract_number": contract_number(row.get("contract_id")), "listing_number": listing_number(row.get("listing_id"))} for row in rows]
    return _create_excel(export_rows, CONTRACT_EXPORT_COLUMNS, "계약 목록")


def create_consultation_excel(rows: list[dict[str, Any]]) -> bytes:
    """상담 조회 결과를 내부 업무용으로 내보낸다. 고객 연락처는 포함하고 이름·자유 메모는 제외한다."""
    export_rows = [{**row, "consultation_number": consultation_number(row.get("consultation_id")), "listing_number": listing_number(row.get("listing_id"))} for row in rows]
    return _create_excel(export_rows, CONSULTATION_EXPORT_COLUMNS, "상담 목록")


def create_today_tasks_excel(tasks: dict[str, list[dict[str, Any]]]) -> bytes:
    """오늘·지연·상시 확인 필요 업무만 명시한 열로 내보낸다."""
    rows = []
    for group in ("오늘", "지연", "상시 확인 필요"):
        for task in tasks.get(group, []):
            rows.append({"구분": group, **{key: value for key, value in task.items() if key != "kind"}})
    return _create_excel(rows, TODAY_TASK_EXPORT_COLUMNS, "오늘 할 일")


def make_export_filename(received_start: str | None, received_end: str | None) -> str:
    """접수일 범위와 생성 시각을 알 수 있는 다운로드 파일 이름을 만든다."""
    start = received_start or "처음"
    end = received_end or "전체"
    created = datetime.now().strftime("%Y%m%d_%H%M")
    return f"매물목록_접수일_{start}_{end}_{created}.xlsx"


def make_management_export_filename(kind: str) -> str:
    """계약·상담 조회 결과의 생성 시각이 드러나는 파일 이름을 만든다."""
    created = datetime.now().strftime("%Y%m%d_%H%M")
    return f"{kind}_조회결과_{created}.xlsx"


def make_today_tasks_export_filename(reference_date: str) -> str:
    """선택한 업무 기준일과 생성 시각이 보이는 파일 이름을 만든다."""
    created = datetime.now().strftime("%Y%m%d_%H%M")
    return f"오늘할일_기준일_{reference_date}_{created}.xlsx"
